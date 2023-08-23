from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.files.storage import FileSystemStorage, default_storage
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse, Http404
from .models import Pegawai, Pendidikan, Jabatan, Pangkat, AngkaKredit, Diklat
from openpyxl import load_workbook
import os
import json
import openpyxl
import re


def index(request):
    # Mendapatkan username user yang sedang login
    username = request.user.username

    ##### PEGAWAI #####
    # Mendapatkan data pegawai dari user yang sedang login
    try:
        # Mengambil data tabel pegawai
        pegawai = Pegawai.objects.get(user=request.user)
        # Ubah format tanggal lahir menjadi "YYYY/MM/DD"
        tanggal_lahir_pegawai = pegawai.tanggal_lahir_pegawai.strftime("%Y-%m-%d")
    except Pegawai.DoesNotExist:
        # Jika data tabel pegawai tidak ada maka return Null/None
        pegawai = None
        tanggal_lahir_pegawai = None

    ##### PENDIDIKAN #####
    # Mendapatkan data pendidikan dari user yang sedang login
    try:
        # Mengambil data tabel pendidikan
        pendidikan = Pendidikan.objects.filter(user=request.user).order_by(
            "-tanggal_terbit_ijazah_pendidikan"
        )
        if pendidikan.exists():
            pendidikan_terakhir = pendidikan[0]
        else:
            pendidikan_terakhir = None
    except Pendidikan.DoesNotExist:
        # Jika data tabel pendidikan tidak ada maka return Null/None
        pendidikan = None
        pendidikan_terakhir = None

    # Inisialisasi file_exists_pendidikan sebagai False
    file_exists_pendidikan = False

    if pendidikan.exists():
        # Ubah format tanggal terbit ijazah pendidikan menjadi "YYYY/MM/DD"
        for pend in pendidikan:
            pend.tanggal_terbit_ijazah_pendidikan = (
                pend.tanggal_terbit_ijazah_pendidikan.strftime("%Y-%m-%d")
            )
            if pend and pend.file_ijazah_pendidikan:
                file_exists_pendidikan = default_storage.exists(
                    pend.file_ijazah_pendidikan.name
                )
            else:
                # Keluar dari loop jika file_exists_pendidikan telah ditemukan
                break

    # Mengambil data pilihan tingkat pendidikan
    pilihan_tingkat_pendidikan = Pendidikan.TingkatPendidikan.choices

    ##### JABATAN #####
    # Mendapatkan data jabatan dari user yang sedang login
    try:
        # Mengambil data tabel jabatan
        jabatan = Jabatan.objects.filter(user=request.user).order_by(
            "-tanggal_sk_jabatan"
        )
        if jabatan.exists():
            jabatan_terakhir = jabatan[0]
        else:
            jabatan_terakhir = None
    except Jabatan.DoesNotExist:
        # Jika data tabel jabatan tidak ada maka return Null/None
        jabatan = None
        jabatan_terakhir = None

    # Inisialisasi file_exists_jabatan sebagai False
    file_exists_jabatan = False

    if jabatan.exists():
        # Ubah format tanggal sk jabatan menjadi "YYYY/MM/DD"
        for jab in jabatan:
            jab.tanggal_sk_jabatan = jab.tanggal_sk_jabatan.strftime("%Y-%m-%d")
            jab.tmt_jabatan = jab.tmt_jabatan.strftime("%Y-%m-%d")
            if jab and jab.file_sk_jabatan:
                file_exists_jabatan = default_storage.exists(jab.file_sk_jabatan.name)
            else:
                # Keluar dari loop jika file_exists_jabatan telah ditemukan
                break

    # Mengambil data pilihan jabatan
    pilihan_jabatan = Jabatan.NamaJabatan.choices

    ##### PANGKAT #####
    # Mendapatkan data pangkat dari user yang sedang login
    try:
        # Mengambil data tabel pangkat
        pangkat = Pangkat.objects.filter(user=request.user).order_by(
            "-tanggal_sk_pangkat"
        )
        if pangkat.exists():
            pangkat_terakhir = pangkat[0]
        else:
            pangkat_terakhir = None
    except Pangkat.DoesNotExist:
        # Jika data tabel pangkat tidak ada maka return Null/None
        pangkat = None
        pangkat_terakhir = None

    # Inisialisasi file_exists_pangkat sebagai False
    file_exists_pangkat = False

    if pangkat.exists():
        # Ubah format tanggal sk pangkat menjadi "YYYY/MM/DD"
        for pang in pangkat:
            pang.tanggal_sk_pangkat = pang.tanggal_sk_pangkat.strftime("%Y-%m-%d")
            pang.tmt_pangkat = pang.tmt_pangkat.strftime("%Y-%m-%d")
            if pang and pang.tanggal_sk_pangkat:
                file_exists_pangkat = default_storage.exists(pang.file_sk_pangkat.name)
            else:
                # Keluar dari loop jika file_exists_pangkat telah ditemukan
                break

    # Mengambil data pilihan pangkat
    pilihan_pangkat = Pangkat.KodePangkat.choices

    ##### PENILAIAN ANGKA KREDIT (PAK) #####
    # Mendapatkan data PAK dari user yang sedang login
    try:
        # Mengambil data tabel PAK
        pak = AngkaKredit.objects.filter(user=request.user).order_by("-tanggal_pak")
        if pak.exists():
            pak_terakhir = pak[0]
        else:
            pak_terakhir = None
    except AngkaKredit.DoesNotExist:
        # Jika data tabel PAK tidak ada maka return Null/None
        pak = None
        pak_terakhir = None

    # Inisialisasi file_exists_pak sebagai False
    file_exists_pak = False

    # Mengambil data PAK untuk dimasukkan ke dalam chart
    pak_data = []

    if pak.exists():
        # Ubah format tanggal PAK menjadi "YYYY/MM/DD"
        for ak in pak:
            ak.tanggal_pak = ak.tanggal_pak.strftime("%Y-%m-%d")
            pak_data.append(
                {
                    "tanggal_pak": ak.tanggal_pak,
                    "nilai_pak": ak.nilai_pak,
                }
            )
            if ak and ak.tanggal_pak:
                file_exists_pak = default_storage.exists(ak.file_pak.name)
            else:
                # Keluar dari loop jika file_exists_pak telah ditemukan
                break

    # Mengubah data menjadi JSON
    pak_data_json = json.dumps(pak_data)

    ##### DIKLAT #####
    # Mendapatkan data DIKLAT dari user yang sedang login
    try:
        # Mengambil data tabel DIKLAT
        diklat = Diklat.objects.filter(user=request.user).order_by(
            "-tanggal_sertifikat_diklat"
        )
        if diklat.exists():
            diklat_terakhir = diklat[0]
        else:
            diklat_terakhir = None
    except Diklat.DoesNotExist:
        # Jika data tabel DIKLAT tidak ada maka return Null/None
        diklat = None
        diklat_terakhir = None

    # Inisialisasi file_exists_dik sebagai False
    file_exists_diklat = False

    if diklat.exists():
        # Ubah format tanggal sertifikat DIKLAT, tanggal mulai dan selesai DIKLAT menjadi "YYYY/MM/DD"
        for dik in diklat:
            dik.tanggal_mulai_diklat = dik.tanggal_mulai_diklat.strftime("%Y-%m-%d")
            dik.tanggal_selesai_diklat = dik.tanggal_selesai_diklat.strftime("%Y-%m-%d")
            dik.tanggal_sertifikat_diklat = dik.tanggal_sertifikat_diklat.strftime(
                "%Y-%m-%d"
            )
            if dik and dik.tanggal_sertifikat_diklat:
                file_exists_diklat = default_storage.exists(
                    dik.file_sertifikat_diklat.name
                )
            else:
                # Keluar dari loop jika file_exists_diklat telah ditemukan
                break

    context = {
        "username": username,
        "pegawai": pegawai,
        "tanggal_lahir_pegawai": tanggal_lahir_pegawai,
        "pendidikan": pendidikan,
        "pendidikan_terakhir": pendidikan_terakhir,
        "pilihan_tingkat_pendidikan": pilihan_tingkat_pendidikan,
        "file_exists_pendidikan": file_exists_pendidikan,
        "jabatan": jabatan,
        "jabatan_terakhir": jabatan_terakhir,
        "pilihan_jabatan": pilihan_jabatan,
        "file_exists_jabatan": file_exists_jabatan,
        "pangkat": pangkat,
        "pangkat_terakhir": pangkat_terakhir,
        "pilihan_pangkat": pilihan_pangkat,
        "file_exists_pangkat": file_exists_pangkat,
        "pak": pak,
        "pak_terakhir": pak_terakhir,
        "file_exists_pak": file_exists_pak,
        "pak_data_json": pak_data_json,  # Menambahkan data JSON ke konteks
        "diklat": diklat,
        "diklat_terakhir": diklat_terakhir,
        "file_exists_diklat": file_exists_diklat,
    }
    return render(request, "pegawai/index.html", context)


@login_required
def profile(request):
    # Mendapatkan username user yang sedang login
    username = request.user.username

    ##### PEGAWAI #####
    # Mendapatkan data pegawai dari user yang sedang login
    try:
        # Mengambil data tabel pegawai
        pegawai = Pegawai.objects.get(user=request.user)
        # Ubah format tanggal lahir menjadi "YYYY/MM/DD"
        tanggal_lahir_pegawai = pegawai.tanggal_lahir_pegawai.strftime("%Y-%m-%d")
    except Pegawai.DoesNotExist:
        # Jika data tabel pegawai tidak ada maka return Null/None
        pegawai = None
        tanggal_lahir_pegawai = None

    ##### PENDIDIKAN #####
    # Mendapatkan data pendidikan dari user yang sedang login
    try:
        # Mengambil data tabel pendidikan
        pendidikan = Pendidikan.objects.filter(user=request.user).order_by(
            "-tanggal_terbit_ijazah_pendidikan"
        )
        if pendidikan.exists():
            pendidikan_terakhir = pendidikan[0]
        else:
            pendidikan_terakhir = None
    except Pendidikan.DoesNotExist:
        # Jika data tabel pendidikan tidak ada maka return Null/None
        pendidikan = None
        pendidikan_terakhir = None

    # Inisialisasi file_exists_pendidikan sebagai False
    file_exists_pendidikan = False

    if pendidikan.exists():
        # Ubah format tanggal terbit ijazah pendidikan menjadi "YYYY/MM/DD"
        for pend in pendidikan:
            pend.tanggal_terbit_ijazah_pendidikan = (
                pend.tanggal_terbit_ijazah_pendidikan.strftime("%Y-%m-%d")
            )
            if pend and pend.file_ijazah_pendidikan:
                file_exists_pendidikan = default_storage.exists(
                    pend.file_ijazah_pendidikan.name
                )
            else:
                # Keluar dari loop jika file_exists_pendidikan telah ditemukan
                break

    # Mengambil data pilihan tingkat pendidikan
    pilihan_tingkat_pendidikan = Pendidikan.TingkatPendidikan.choices

    ##### JABATAN #####
    # Mendapatkan data jabatan dari user yang sedang login
    try:
        # Mengambil data tabel jabatan
        jabatan = Jabatan.objects.filter(user=request.user).order_by(
            "-tanggal_sk_jabatan"
        )
        if jabatan.exists():
            jabatan_terakhir = jabatan[0]
        else:
            jabatan_terakhir = None
    except Jabatan.DoesNotExist:
        # Jika data tabel jabatan tidak ada maka return Null/None
        jabatan = None
        jabatan_terakhir = None

    # Inisialisasi file_exists_jabatan sebagai False
    file_exists_jabatan = False

    if jabatan.exists():
        # Ubah format tanggal sk jabatan menjadi "YYYY/MM/DD"
        for jab in jabatan:
            jab.tanggal_sk_jabatan = jab.tanggal_sk_jabatan.strftime("%Y-%m-%d")
            jab.tmt_jabatan = jab.tmt_jabatan.strftime("%Y-%m-%d")
            if jab and jab.file_sk_jabatan:
                file_exists_jabatan = default_storage.exists(jab.file_sk_jabatan.name)
            else:
                # Keluar dari loop jika file_exists_jabatan telah ditemukan
                break

    # Mengambil data pilihan jabatan
    pilihan_jabatan = Jabatan.NamaJabatan.choices

    ##### PANGKAT #####
    # Mendapatkan data pangkat dari user yang sedang login
    try:
        # Mengambil data tabel pangkat
        pangkat = Pangkat.objects.filter(user=request.user).order_by(
            "-tanggal_sk_pangkat"
        )
        if pangkat.exists():
            pangkat_terakhir = pangkat[0]
        else:
            pangkat_terakhir = None
    except Pangkat.DoesNotExist:
        # Jika data tabel pangkat tidak ada maka return Null/None
        pangkat = None
        pangkat_terakhir = None

    # Inisialisasi file_exists_pangkat sebagai False
    file_exists_pangkat = False

    if pangkat.exists():
        # Ubah format tanggal sk pangkat menjadi "YYYY/MM/DD"
        for pang in pangkat:
            pang.tanggal_sk_pangkat = pang.tanggal_sk_pangkat.strftime("%Y-%m-%d")
            pang.tmt_pangkat = pang.tmt_pangkat.strftime("%Y-%m-%d")
            if pang and pang.tanggal_sk_pangkat:
                file_exists_pangkat = default_storage.exists(pang.file_sk_pangkat.name)
            else:
                # Keluar dari loop jika file_exists_pangkat telah ditemukan
                break

    # Mengambil data pilihan pangkat
    pilihan_pangkat = Pangkat.KodePangkat.choices

    ##### PENILAIAN ANGKA KREDIT (PAK) #####
    # Mendapatkan data PAK dari user yang sedang login
    try:
        # Mengambil data tabel PAK
        pak = AngkaKredit.objects.filter(user=request.user).order_by("-tanggal_pak")
        if pak.exists():
            pak_terakhir = pak[0]
        else:
            pak_terakhir = None
    except AngkaKredit.DoesNotExist:
        # Jika data tabel PAK tidak ada maka return Null/None
        pak = None
        pak_terakhir = None

    # Inisialisasi file_exists_pak sebagai False
    file_exists_pak = False

    # Mengambil data PAK untuk dimasukkan ke dalam chart
    pak_data = []

    if pak.exists():
        # Ubah format tanggal PAK menjadi "YYYY/MM/DD"
        for ak in pak:
            ak.tanggal_pak = ak.tanggal_pak.strftime("%Y-%m-%d")
            pak_data.append(
                {
                    "tanggal_pak": ak.tanggal_pak,
                    "nilai_pak": ak.nilai_pak,
                }
            )
            if ak and ak.tanggal_pak:
                file_exists_pak = default_storage.exists(ak.file_pak.name)
            else:
                # Keluar dari loop jika file_exists_pak telah ditemukan
                break

    # Mengubah data menjadi JSON
    pak_data_json = json.dumps(pak_data)

    ##### DIKLAT #####
    # Mendapatkan data DIKLAT dari user yang sedang login
    try:
        # Mengambil data tabel DIKLAT
        diklat = Diklat.objects.filter(user=request.user).order_by(
            "-tanggal_sertifikat_diklat"
        )
        if diklat.exists():
            diklat_terakhir = diklat[0]
        else:
            diklat_terakhir = None
    except Diklat.DoesNotExist:
        # Jika data tabel DIKLAT tidak ada maka return Null/None
        diklat = None
        diklat_terakhir = None

    # Inisialisasi file_exists_dik sebagai False
    file_exists_diklat = False

    if diklat.exists():
        # Ubah format tanggal sertifikat DIKLAT, tanggal mulai dan selesai DIKLAT menjadi "YYYY/MM/DD"
        for dik in diklat:
            dik.tanggal_mulai_diklat = dik.tanggal_mulai_diklat.strftime("%Y-%m-%d")
            dik.tanggal_selesai_diklat = dik.tanggal_selesai_diklat.strftime("%Y-%m-%d")
            dik.tanggal_sertifikat_diklat = dik.tanggal_sertifikat_diklat.strftime(
                "%Y-%m-%d"
            )
            if dik and dik.tanggal_sertifikat_diklat:
                file_exists_diklat = default_storage.exists(
                    dik.file_sertifikat_diklat.name
                )
            else:
                # Keluar dari loop jika file_exists_diklat telah ditemukan
                break

    context = {
        "username": username,
        "pegawai": pegawai,
        "tanggal_lahir_pegawai": tanggal_lahir_pegawai,
        "pendidikan": pendidikan,
        "pendidikan_terakhir": pendidikan_terakhir,
        "pilihan_tingkat_pendidikan": pilihan_tingkat_pendidikan,
        "file_exists_pendidikan": file_exists_pendidikan,
        "jabatan": jabatan,
        "jabatan_terakhir": jabatan_terakhir,
        "pilihan_jabatan": pilihan_jabatan,
        "file_exists_jabatan": file_exists_jabatan,
        "pangkat": pangkat,
        "pangkat_terakhir": pangkat_terakhir,
        "pilihan_pangkat": pilihan_pangkat,
        "file_exists_pangkat": file_exists_pangkat,
        "pak": pak,
        "pak_terakhir": pak_terakhir,
        "file_exists_pak": file_exists_pak,
        "pak_data_json": pak_data_json,  # Menambahkan data JSON ke konteks
        "diklat": diklat,
        "diklat_terakhir": diklat_terakhir,
        "file_exists_diklat": file_exists_diklat,
    }
    return render(request, "pegawai/profile.html", context)


def signin(request):
    if request.user.is_authenticated:
        return redirect("index")  # User sudah signin, arahkan ke index

    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("index")
        else:
            messages.error(request, "Username atau password salah.")

    return render(request, "pegawai/signin.html")


@login_required
def signout(request):
    logout(request)
    return redirect("signin")


def signup(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        confirm_password = request.POST["confirm_password"]

        if not re.match("^[a-zA-Z0-9]+$", username):
            messages.error(
                request, "Username hanya boleh terdiri dari huruf dan angka."
            )
        elif password == confirm_password:
            if User.objects.filter(username=username).exists():
                messages.error(
                    request, "Username sudah terdaftar. Silakan gunakan username lain."
                )
            else:
                # Buat user baru
                User.objects.create_user(username=username, password=password)
                messages.success(request, "Anda berhasil mendaftar, silahkan login.")
                return redirect("signin")
            # Ganti 'home' dengan URL halaman setelah sign up
        else:
            messages.error(request, "Password tidak sesuai.")

    return render(request, "pegawai/signup.html")


##### PEGAWAI #####
@login_required
def add_pegawai(request):
    user_id = request.user.id

    if request.method == "POST":
        aktif_pegawai = request.POST["aktif_pegawai"]
        nip_pegawai = request.POST["nip_pegawai"]
        nama_pegawai = request.POST["nama_pegawai"]
        tempat_lahir_pegawai = request.POST["tempat_lahir_pegawai"]
        tanggal_lahir_pegawai = request.POST["tanggal_lahir_pegawai"]
        jenis_kelamin_pegawai = request.POST["jenis_kelamin_pegawai"]
        surel_pegawai = request.POST["surel_pegawai"]
        telepon_pegawai = request.POST["telepon_pegawai"]

        # Cek apakah sudah ada pegawai dengan user_id yang sama
        try:
            pegawai = Pegawai.objects.get(user_id=user_id)
            # Update data pegawai
            pegawai.aktif_pegawai = aktif_pegawai
            pegawai.nip_pegawai = nip_pegawai
            pegawai.nama_pegawai = nama_pegawai
            pegawai.tempat_lahir_pegawai = tempat_lahir_pegawai
            pegawai.tanggal_lahir_pegawai = tanggal_lahir_pegawai
            pegawai.jenis_kelamin_pegawai = jenis_kelamin_pegawai
            pegawai.surel_pegawai = surel_pegawai
            pegawai.telepon_pegawai = telepon_pegawai
            pegawai.save()
            messages.success(request, "Data pegawai berhasil diperbarui.")
        except Pegawai.DoesNotExist:
            # Buat pegawai baru
            Pegawai.objects.create(
                aktif_pegawai=aktif_pegawai,
                nip_pegawai=nip_pegawai,
                nama_pegawai=nama_pegawai,
                tempat_lahir_pegawai=tempat_lahir_pegawai,
                tanggal_lahir_pegawai=tanggal_lahir_pegawai,
                jenis_kelamin_pegawai=jenis_kelamin_pegawai,
                surel_pegawai=surel_pegawai,
                telepon_pegawai=telepon_pegawai,
                user_id=user_id,
            )
            messages.success(request, "Pegawai berhasil ditambahkan.")

        return redirect("profile")

    return render(request, "pegawai/profile.html")


@login_required
@user_passes_test(lambda user: user.is_superuser or user.is_staff)
def add_data_pegawai(request):
    if request.method == "POST":
        aktif_pegawai = request.POST["aktif_pegawai"]
        nip_pegawai = request.POST["nip_pegawai"]
        nama_pegawai = request.POST["nama_pegawai"]
        tempat_lahir_pegawai = request.POST["tempat_lahir_pegawai"]
        tanggal_lahir_pegawai = request.POST["tanggal_lahir_pegawai"]
        jenis_kelamin_pegawai = request.POST["jenis_kelamin_pegawai"]
        surel_pegawai = request.POST["surel_pegawai"]
        telepon_pegawai = request.POST["telepon_pegawai"]
        user = request.POST["user"]

        # Cek apakah sudah ada pegawai dengan user_id yang sama
        try:
            pegawai = Pegawai.objects.get(user_id=user)
            # Update data pegawai
            pegawai.aktif_pegawai = aktif_pegawai
            pegawai.nip_pegawai = nip_pegawai
            pegawai.nama_pegawai = nama_pegawai
            pegawai.tempat_lahir_pegawai = tempat_lahir_pegawai
            pegawai.tanggal_lahir_pegawai = tanggal_lahir_pegawai
            pegawai.jenis_kelamin_pegawai = jenis_kelamin_pegawai
            pegawai.surel_pegawai = surel_pegawai
            pegawai.telepon_pegawai = telepon_pegawai
            pegawai.user = user
            pegawai.save()
            messages.success(request, "Data pegawai berhasil diperbarui.")
        except Pegawai.DoesNotExist:
            # Buat pegawai baru
            Pegawai.objects.create(
                aktif_pegawai=aktif_pegawai,
                nip_pegawai=nip_pegawai,
                nama_pegawai=nama_pegawai,
                tempat_lahir_pegawai=tempat_lahir_pegawai,
                tanggal_lahir_pegawai=tanggal_lahir_pegawai,
                jenis_kelamin_pegawai=jenis_kelamin_pegawai,
                surel_pegawai=surel_pegawai,
                telepon_pegawai=telepon_pegawai,
                user_id=user,
            )
            messages.success(request, "Pegawai berhasil ditambahkan.")

        return redirect("administrasi")

    return redirect("administrasi")


@login_required
@user_passes_test(lambda user: user.is_superuser or user.is_staff)
def edit_pegawai(request, id_pegawai):
    try:
        pegawai = get_object_or_404(Pegawai, id_pegawai=id_pegawai)
    except Pegawai.DoesNotExist:
        messages.error(request, "Data pegawai tidak ditemukan.")
        return redirect("administrasi")

    if request.method == "POST":
        aktif_pegawai = request.POST["aktif_pegawai"]
        nip_pegawai = request.POST["nip_pegawai"]
        nama_pegawai = request.POST["nama_pegawai"]
        tempat_lahir_pegawai = request.POST["tempat_lahir_pegawai"]
        tanggal_lahir_pegawai = request.POST["tanggal_lahir_pegawai"]
        jenis_kelamin_pegawai = request.POST["jenis_kelamin_pegawai"]
        surel_pegawai = request.POST["surel_pegawai"]
        telepon_pegawai = request.POST["telepon_pegawai"]
        user = request.POST["user"]

        # Update data pegawai
        pegawai.aktif_pegawai = aktif_pegawai
        pegawai.nip_pegawai = nip_pegawai
        pegawai.nama_pegawai = nama_pegawai
        pegawai.tempat_lahir_pegawai = tempat_lahir_pegawai
        pegawai.tanggal_lahir_pegawai = tanggal_lahir_pegawai
        pegawai.jenis_kelamin_pegawai = jenis_kelamin_pegawai
        pegawai.surel_pegawai = surel_pegawai
        pegawai.telepon_pegawai = telepon_pegawai
        pegawai.user = user
        pegawai.save()

        messages.success(request, "Data Pegawai berhasil diperbarui.")

        return redirect("administrasi")

    return redirect("administrasi")


@login_required
@user_passes_test(lambda user: user.is_superuser or user.is_staff)
def delete_pegawai(request, id_pegawai):
    pegawai = get_object_or_404(Pegawai, id_pegawai=id_pegawai)
    try:
        pegawai.delete()
        messages.success(request, "Data Pegawai berhasil dihapus.")
    except Pangkat.DoesNotExist:
        messages.warning(request, "Data Pegawai tidak ditemukan.")

    return redirect("administrasi")


##### PENDIDIKAN #####
@login_required
def add_pendidikan(request):
    user_id = request.user.id

    if request.method == "POST" and request.FILES["file_ijazah_pendidikan"]:
        tingkat_pendidikan = request.POST["tingkat_pendidikan"]
        lembaga_pendidikan = request.POST["lembaga_pendidikan"]
        fakultas_pendidikan = request.POST["fakultas_pendidikan"]
        jurusan_pendidikan = request.POST["jurusan_pendidikan"]
        gelar_depan_pendidikan = request.POST["gelar_depan_pendidikan"]
        gelar_belakang_pendidikan = request.POST["gelar_belakang_pendidikan"]
        nomor_seri_ijazah_pendidikan = request.POST["nomor_seri_ijazah_pendidikan"]
        tanggal_terbit_ijazah_pendidikan = request.POST[
            "tanggal_terbit_ijazah_pendidikan"
        ]
        file_ijazah_pendidikan = request.FILES["file_ijazah_pendidikan"]

        # Menggunakan get untuk mendapatkan nilai, jika ada
        user_pendidikan = request.POST.get("user")

        # Validasi tipe file
        if not file_ijazah_pendidikan.name.endswith(".pdf"):
            messages.error(request, "File harus berformat PDF.")
            return redirect("profile")

        # Validasi ukuran file
        max_size = 1000 * 1024  # 1000KB
        if file_ijazah_pendidikan.size > max_size:
            messages.error(request, "Ukuran file melebihi 1000KB.")
            return redirect("profile")

        # Dapatkan username berdasarkan user_pendidikan (jika ada) atau user saat ini
        if user_pendidikan is not None:
            # Cari objek User berdasarkan id_user
            username = User.objects.get(id=user_pendidikan).username
        else:
            # Dapatkan username dari pengguna yang saat ini masuk
            username = request.user.username

        # Ubah nama file yang akan diunggah
        new_file_name = f"{username}_pendidikan_{tingkat_pendidikan}{file_ijazah_pendidikan.name[file_ijazah_pendidikan.name.rfind('.'):]}"

        fs = FileSystemStorage()

        try:
            # Cek apakah data pendidikan sudah ada
            Pendidikan.objects.get(
                tingkat_pendidikan=tingkat_pendidikan,
                nomor_seri_ijazah_pendidikan=nomor_seri_ijazah_pendidikan,
                user_id=user_pendidikan if user_pendidikan is not None else user_id,
            )
            messages.warning(
                request,
                "Data Pendidikan sudah ada. Tolong gunakan fungsi edit.",
            )
        except Pendidikan.DoesNotExist:
            # Buat data pendidikan baru
            Pendidikan.objects.create(
                tingkat_pendidikan=tingkat_pendidikan,
                lembaga_pendidikan=lembaga_pendidikan,
                fakultas_pendidikan=fakultas_pendidikan,
                jurusan_pendidikan=jurusan_pendidikan,
                gelar_depan_pendidikan=gelar_depan_pendidikan,
                gelar_belakang_pendidikan=gelar_belakang_pendidikan,
                nomor_seri_ijazah_pendidikan=nomor_seri_ijazah_pendidikan,
                tanggal_terbit_ijazah_pendidikan=tanggal_terbit_ijazah_pendidikan,
                file_ijazah_pendidikan=fs.save(
                    f"media/pendidikan/{new_file_name}", file_ijazah_pendidikan
                ),
                user_id=user_pendidikan if user_pendidikan is not None else user_id,
            )
            messages.success(request, "Data Pendidikan berhasil dibuat.")

        return redirect("administrasi" if user_pendidikan is not None else "profile")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def edit_pendidikan(request, id_pendidikan):
    user_id = request.user.id

    if request.method == "POST" and request.FILES["file_ijazah_pendidikan"]:
        tingkat_pendidikan = request.POST["tingkat_pendidikan"]
        lembaga_pendidikan = request.POST["lembaga_pendidikan"]
        fakultas_pendidikan = request.POST["fakultas_pendidikan"]
        jurusan_pendidikan = request.POST["jurusan_pendidikan"]
        gelar_depan_pendidikan = request.POST["gelar_depan_pendidikan"]
        gelar_belakang_pendidikan = request.POST["gelar_belakang_pendidikan"]
        nomor_seri_ijazah_pendidikan = request.POST["nomor_seri_ijazah_pendidikan"]
        tanggal_terbit_ijazah_pendidikan = request.POST[
            "tanggal_terbit_ijazah_pendidikan"
        ]
        file_ijazah_pendidikan = request.FILES["file_ijazah_pendidikan"]

        # Menggunakan get untuk mendapatkan nilai, jika ada
        user_pendidikan = request.POST.get("user")

        # Validasi tipe file
        if not file_ijazah_pendidikan.name.endswith(".pdf"):
            messages.error(request, "File harus berformat PDF.")
            return redirect("profile")

        # Validasi ukuran file
        max_size = 1000 * 1024  # 1000KB
        if file_ijazah_pendidikan.size > max_size:
            messages.error(request, "Ukuran file melebihi 1000KB.")
            return redirect("profile")

        # Dapatkan username berdasarkan user_pendidikan (jika ada) atau user saat ini
        if user_pendidikan is not None:
            # Cari objek User berdasarkan id_user
            username = User.objects.get(id=user_pendidikan).username
        else:
            # Dapatkan username dari pengguna yang saat ini masuk
            username = request.user.username

        # Ubah nama file yang akan diunggah
        new_file_name = f"{username}_pendidikan_{tingkat_pendidikan}{file_ijazah_pendidikan.name[file_ijazah_pendidikan.name.rfind('.'):]}"

        fs = FileSystemStorage()

        try:
            # Variabel user tidak memiliki nilai dari form
            # Lakukan tindakan alternatif
            pendidikan = get_object_or_404(
                Pendidikan,
                id_pendidikan=id_pendidikan,
                user=user_pendidikan if user_pendidikan is not None else user_id,
            )
            # Hapus file dan data pendidikan sebelumnya
            if pendidikan.file_ijazah_pendidikan:
                old_file_path = pendidikan.file_ijazah_pendidikan.path
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)

            # Update data pendidikan
            pendidikan.tingkat_pendidikan = tingkat_pendidikan
            pendidikan.lembaga_pendidikan = lembaga_pendidikan
            pendidikan.fakultas_pendidikan = fakultas_pendidikan
            pendidikan.jurusan_pendidikan = jurusan_pendidikan
            pendidikan.gelar_depan_pendidikan = gelar_depan_pendidikan
            pendidikan.gelar_belakang_pendidikan = gelar_belakang_pendidikan
            pendidikan.nomor_seri_ijazah_pendidikan = nomor_seri_ijazah_pendidikan
            pendidikan.tanggal_terbit_ijazah_pendidikan = (
                tanggal_terbit_ijazah_pendidikan
            )
            pendidikan.file_ijazah_pendidikan = fs.save(
                f"media/pendidikan/{new_file_name}", file_ijazah_pendidikan
            )
            pendidikan.save()
            messages.success(request, "Data Pendidikan berhasil diubah.")
        except Pendidikan.DoesNotExist:
            messages.warning(
                request,
                "Data Pendidikan tidak ditemukan. Tolong gunakan fungsi buat.",
            )

        return redirect("administrasi" if user_pendidikan is not None else "profile")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def delete_pendidikan(request, id_pendidikan):
    pendidikan = get_object_or_404(Pendidikan, id_pendidikan=id_pendidikan)

    try:
        # Hapus file dan data pendidikan
        if pendidikan.file_ijazah_pendidikan:
            old_file_path = pendidikan.file_ijazah_pendidikan.path
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        pendidikan.delete()
        messages.success(request, "Data Pendidikan berhasil dihapus.")
    except Pendidikan.DoesNotExist:
        messages.warning(request, "Data Pendidikan tidak ditemukan.")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def download_ijazah(request, id_pendidikan):
    pendidikan = get_object_or_404(Pendidikan, id_pendidikan=id_pendidikan)

    # Pastikan pengguna yang mengakses adalah pemilik data ijazah
    if pendidikan.user != request.user:
        messages.error(request, "Anda tidak memiliki izin untuk mengunduh ijazah ini.")
        return redirect("profile")

    # Cek apakah file_ijazah_pendidikan ada atau tidak
    if not pendidikan.file_ijazah_pendidikan:
        raise Http404("File Ijazah tidak ditemukan")

    # Dapatkan path lengkap ke file ijazah
    file_path = pendidikan.file_ijazah_pendidikan.path

    # Buka file sebagai binary dan kirim sebagai HTTP response
    with open(file_path, "rb") as file:
        response = HttpResponse(
            file.read(), content_type="application/pdf"
        )  # Sesuaikan content_type sesuai tipe file

        # Mengatur header untuk attachment agar file dapat diunduh
        response[
            "Content-Disposition"
        ] = f"attachment; filename={pendidikan.file_ijazah_pendidikan.name}"

        return response


##### JABATAN #####
@login_required
def add_jabatan(request):
    user_id = request.user.id

    if request.method == "POST" and request.FILES["file_sk_jabatan"]:
        nama_jabatan = request.POST["nama_jabatan"]
        nomor_sk_jabatan = request.POST["nomor_sk_jabatan"]
        tanggal_sk_jabatan = request.POST["tanggal_sk_jabatan"]
        tmt_jabatan = request.POST["tmt_jabatan"]
        file_sk_jabatan = request.FILES["file_sk_jabatan"]

        # Menggunakan get untuk mendapatkan nilai, jika ada
        user_jabatan = request.POST.get("user")

        # Validasi tipe file
        if not file_sk_jabatan.name.endswith(".pdf"):
            messages.error(request, "File harus berformat PDF.")
            return redirect("profile")

        # Validasi ukuran file
        max_size = 1000 * 1024  # 1000KB
        if file_sk_jabatan.size > max_size:
            messages.error(request, "Ukuran file melebihi 1000KB.")
            return redirect("profile")

        # Dapatkan username berdasarkan user_jabatan (jika ada) atau user saat ini
        if user_jabatan is not None:
            # Cari objek User berdasarkan id_user
            username = User.objects.get(id=user_jabatan).username
        else:
            # Dapatkan username dari pengguna yang saat ini masuk
            username = request.user.username

        # Ubah nama file yang akan diunggah
        new_file_name = f"{username}_jabatan_{nama_jabatan}{file_sk_jabatan.name[file_sk_jabatan.name.rfind('.'):]}"

        fs = FileSystemStorage()

        try:
            # Cek apakah data jabatan sudah ada
            Jabatan.objects.get(
                nama_jabatan=nama_jabatan,
                nomor_sk_jabatan=nomor_sk_jabatan,
                user_id=user_jabatan if user_jabatan is not None else user_id,
            )
            messages.warning(
                request,
                "Data Jabatan sudah ada. Tolong gunakan fungsi edit.",
            )
            return redirect("profile")
        except Jabatan.DoesNotExist:
            # Buat data jabatan baru
            Jabatan.objects.create(
                nama_jabatan=nama_jabatan,
                nomor_sk_jabatan=nomor_sk_jabatan,
                tanggal_sk_jabatan=tanggal_sk_jabatan,
                tmt_jabatan=tmt_jabatan,
                file_sk_jabatan=fs.save(
                    f"media/jabatan/{new_file_name}", file_sk_jabatan
                ),
                user_id=user_jabatan if user_jabatan is not None else user_id,
            )
            messages.success(request, "Data Jabatan berhasil dibuat.")

        return redirect("administrasi" if user_jabatan is not None else "profile")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def edit_jabatan(request, id_jabatan):
    user_id = request.user.id

    if request.method == "POST" and request.FILES["file_sk_jabatan"]:
        nama_jabatan = request.POST["nama_jabatan"]
        nomor_sk_jabatan = request.POST["nomor_sk_jabatan"]
        tanggal_sk_jabatan = request.POST["tanggal_sk_jabatan"]
        tmt_jabatan = request.POST["tmt_jabatan"]
        file_sk_jabatan = request.FILES["file_sk_jabatan"]

        # Menggunakan get untuk mendapatkan nilai, jika ada
        user_jabatan = request.POST.get("user")

        # Validasi tipe file
        if not file_sk_jabatan.name.endswith(".pdf"):
            messages.error(request, "File harus berformat PDF.")
            return redirect("profile")

        # Validasi ukuran file
        max_size = 1000 * 1024  # 1000KB
        if file_sk_jabatan.size > max_size:
            messages.error(request, "Ukuran file melebihi 1000KB.")
            return redirect("profile")

        # Dapatkan username berdasarkan user_jabatan (jika ada) atau user saat ini
        if user_jabatan is not None:
            # Cari objek User berdasarkan id_user
            username = User.objects.get(id=user_jabatan).username
        else:
            # Dapatkan username dari pengguna yang saat ini masuk
            username = request.user.username

        # Ubah nama file yang akan diunggah
        new_file_name = f"{username}_jabatan_{nama_jabatan}{file_sk_jabatan.name[file_sk_jabatan.name.rfind('.'):]}"

        fs = FileSystemStorage()

        try:
            jabatan = get_object_or_404(
                Jabatan,
                id_jabatan=id_jabatan,
                user=user_jabatan if user_jabatan is not None else user_id,
            )
            # Hapus file dan data jabatan sebelumnya
            if jabatan.file_sk_jabatan:
                old_file_path = jabatan.file_sk_jabatan.path
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)

            # Update data jabatan
            jabatan.nama_jabatan = nama_jabatan
            jabatan.nomor_sk_jabatan = nomor_sk_jabatan
            jabatan.tanggal_sk_jabatan = tanggal_sk_jabatan
            jabatan.tmt_jabatan = tmt_jabatan
            jabatan.file_sk_jabatan = fs.save(
                f"media/jabatan/{new_file_name}", file_sk_jabatan
            )
            jabatan.save()
            messages.success(request, "Data Jabatan berhasil diubah.")
        except Jabatan.DoesNotExist:
            messages.warning(
                request, "Data Jabatan tidak ditemukan. Tolong gunakan fungsi buat."
            )

        return redirect("administrasi" if user_jabatan is not None else "profile")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def delete_jabatan(request, id_jabatan):
    jabatan = get_object_or_404(Jabatan, id_jabatan=id_jabatan)

    try:
        # Hapus file dan data jabatan
        if jabatan.file_sk_jabatan:
            old_file_path = jabatan.file_sk_jabatan.path
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        jabatan.delete()
        messages.success(request, "Data Jabatan berhasil dihapus.")
    except Jabatan.DoesNotExist:
        messages.warning(request, "Data Jabatan tidak ditemukan.")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def download_sk_jabatan(request, id_jabatan):
    jabatan = get_object_or_404(Jabatan, id_jabatan=id_jabatan)

    # Pastikan pengguna yang mengakses adalah pemilik data SK
    if jabatan.user != request.user:
        messages.error(request, "Anda tidak memiliki izin untuk mengunduh SK ini.")
        return redirect("profile")

    # Cek apakah file_sk_jabatan ada atau tidak
    if not jabatan.file_sk_jabatan:
        raise Http404("File SK tidak ditemukan")

    # Dapatkan path lengkap ke file SK
    file_path = jabatan.file_sk_jabatan.path

    # Buka file sebagai binary dan kirim sebagai HTTP response
    with open(file_path, "rb") as file:
        response = HttpResponse(
            file.read(), content_type="application/pdf"
        )  # Sesuaikan content_type sesuai tipe file

        # Mengatur header untuk attachment agar file dapat diunduh
        response[
            "Content-Disposition"
        ] = f"attachment; filename={jabatan.file_sk_jabatan.name}"

        return response


##### PANGKAT DAN GOLONGAN #####
@login_required
def add_pangkat(request):
    user_id = request.user.id

    if request.method == "POST" and request.FILES["file_sk_pangkat"]:
        nama_pangkat = request.POST["nama_pangkat"]
        nomor_sk_pangkat = request.POST["nomor_sk_pangkat"]
        tanggal_sk_pangkat = request.POST["tanggal_sk_pangkat"]
        tmt_pangkat = request.POST["tmt_pangkat"]
        file_sk_pangkat = request.FILES["file_sk_pangkat"]

        # Menggunakan get untuk mendapatkan nilai, jika ada
        user_pangkat = request.POST.get("user")

        # Validasi tipe file
        if not file_sk_pangkat.name.endswith(".pdf"):
            messages.error(request, "File harus berformat PDF.")
            return redirect("profile")

        # Validasi ukuran file
        max_size = 1000 * 1024  # 1000KB
        if file_sk_pangkat.size > max_size:
            messages.error(request, "Ukuran file melebihi 1000KB.")
            return redirect("profile")

        # Dapatkan username berdasarkan user_pangkat (jika ada) atau user saat ini
        if user_pangkat is not None:
            # Cari objek User berdasarkan id_user
            username = User.objects.get(id=user_pangkat).username
        else:
            # Dapatkan username dari pengguna yang saat ini masuk
            username = request.user.username

        # Ubah nama file yang akan diunggah
        new_file_name = f"{username}_pangkat_{nama_pangkat}{file_sk_pangkat.name[file_sk_pangkat.name.rfind('.'):]}"

        fs = FileSystemStorage()

        try:
            Pangkat.objects.get(
                nama_pangkat=nama_pangkat,
                nomor_sk_pangkat=nomor_sk_pangkat,
                user_id=user_pangkat if user_pangkat is not None else user_id,
            )
            messages.warning(
                request,
                "Data Pangkat dan Golongan sudah ada. Tolong gunakan fungsi edit.",
            )
            return redirect("profile")
        except Pangkat.DoesNotExist:
            Pangkat.objects.create(
                nama_pangkat=nama_pangkat,
                nomor_sk_pangkat=nomor_sk_pangkat,
                tanggal_sk_pangkat=tanggal_sk_pangkat,
                tmt_pangkat=tmt_pangkat,
                file_sk_pangkat=fs.save(
                    f"media/pangkat/{new_file_name}", file_sk_pangkat
                ),
                user_id=user_id,
            )
            messages.success(request, "Data Pangkat berhasil dibuat.")

        return redirect("administrasi" if user_pangkat is not None else "profile")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def edit_pangkat(request, id_pangkat):
    user_id = request.user.id

    if request.method == "POST" and request.FILES["file_sk_pangkat"]:
        nama_pangkat = request.POST["nama_pangkat"]
        nomor_sk_pangkat = request.POST["nomor_sk_pangkat"]
        tanggal_sk_pangkat = request.POST["tanggal_sk_pangkat"]
        tmt_pangkat = request.POST["tmt_pangkat"]
        file_sk_pangkat = request.FILES["file_sk_pangkat"]

        # Menggunakan get untuk mendapatkan nilai, jika ada
        user_pangkat = request.POST.get("user")

        # Validasi tipe file
        if not file_sk_pangkat.name.endswith(".pdf"):
            messages.error(request, "File harus berformat PDF.")
            return redirect("profile")

        # Validasi ukuran file
        max_size = 1000 * 1024  # 1000KB
        if file_sk_pangkat.size > max_size:
            messages.error(request, "Ukuran file melebihi 1000KB.")
            return redirect("profile")

        # Dapatkan username berdasarkan user_pangkat (jika ada) atau user saat ini
        if user_pangkat is not None:
            # Cari objek User berdasarkan id_user
            username = User.objects.get(id=user_pangkat).username
        else:
            # Dapatkan username dari pengguna yang saat ini masuk
            username = request.user.username

        # Ubah nama file yang akan diunggah
        new_file_name = f"{username}_pangkat_{nama_pangkat}{file_sk_pangkat.name[file_sk_pangkat.name.rfind('.'):]}"

        fs = FileSystemStorage()

        try:
            pangkat = get_object_or_404(
                Pangkat,
                id_pangkat=id_pangkat,
                user=user_pangkat if user_pangkat is not None else user_id,
            )
            # Hapus file dan data pangkat sebelumnya
            if pangkat.file_sk_pangkat:
                old_file_path = pangkat.file_sk_pangkat.path
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)

            # Update data pangkat
            pangkat.nama_pangkat = nama_pangkat
            pangkat.nomor_sk_pangkat = nomor_sk_pangkat
            pangkat.tanggal_sk_pangkat = tanggal_sk_pangkat
            pangkat.tmt_pangkat = tmt_pangkat
            pangkat.file_sk_pangkat = fs.save(
                f"media/pangkat/{new_file_name}", file_sk_pangkat
            )
            pangkat.save()
            messages.success(request, "Data Pangkat dan Golongan berhasil diubah.")
        except Pangkat.DoesNotExist:
            messages.warning(
                request, "Data Pangkat tidak ditemukan. Tolong gunakan fungsi buat."
            )

        return redirect("administrasi" if user_pangkat is not None else "profile")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def delete_pangkat(request, id_pangkat):
    pangkat = get_object_or_404(Pangkat, id_pangkat=id_pangkat)

    try:
        # Hapus file dan data pangkat
        if pangkat.file_sk_pangkat:
            old_file_path = pangkat.file_sk_pangkat.path
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        pangkat.delete()
        messages.success(request, "Data Pangkat dan Golongan berhasil dihapus.")
    except Pangkat.DoesNotExist:
        messages.warning(request, "Data Pangkat dan Golongan tidak ditemukan.")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def download_sk_pangkat(request, id_pangkat):
    pangkat = get_object_or_404(Pangkat, id_pangkat=id_pangkat)

    # Pastikan pengguna yang mengakses adalah pemilik data SK
    if pangkat.user != request.user:
        messages.error(request, "Anda tidak memiliki izin untuk mengunduh SK ini.")
        return redirect("profile")

    # Cek apakah file_sk_pangkat ada atau tidak
    if not pangkat.file_sk_pangkat:
        raise Http404("File SK tidak ditemukan")

    # Dapatkan path lengkap ke file SK
    file_path = pangkat.file_sk_pangkat.path

    # Buka file sebagai binary dan kirim sebagai HTTP response
    with open(file_path, "rb") as file:
        response = HttpResponse(
            file.read(), content_type="application/pdf"
        )  # Sesuaikan content_type sesuai tipe file

        # Mengatur header untuk attachment agar file dapat diunduh
        response[
            "Content-Disposition"
        ] = f"attachment; filename={pangkat.file_sk_pangkat.name}"

        return response


##### PENILAIAN ANGKA KREDIT (PAK) #####
@login_required
def add_pak(request):
    user_id = request.user.id

    if request.method == "POST" and request.FILES["file_pak"]:
        nomor_pak = request.POST["nomor_pak"]
        tanggal_pak = request.POST["tanggal_pak"]
        nilai_pak = request.POST["nilai_pak"]
        masa_penilaian_pak = request.POST["masa_penilaian_pak"]
        file_pak = request.FILES["file_pak"]

        # Menggunakan get untuk mendapatkan nilai, jika ada
        user_pak = request.POST.get("user")

        # Validasi tipe file
        if not file_pak.name.endswith(".pdf"):
            messages.error(request, "File harus berformat PDF.")
            return redirect("profile")

        # Validasi ukuran file
        max_size = 1000 * 1024  # 1000KB
        if file_pak.size > max_size:
            messages.error(request, "Ukuran file melebihi 1000KB.")
            return redirect("profile")

        # Dapatkan username berdasarkan user_pak (jika ada) atau user saat ini
        if user_pak is not None:
            # Cari objek User berdasarkan id_user
            username = User.objects.get(id=user_pak).username
        else:
            # Dapatkan username dari pengguna yang saat ini masuk
            username = request.user.username

        # Ubah nama file yang akan diunggah
        new_file_name = (
            f"{username}_pak_{nomor_pak}{file_pak.name[file_pak.name.rfind('.'):]}"
        )

        fs = FileSystemStorage()

        try:
            AngkaKredit.objects.get(
                nomor_pak=nomor_pak,
                tanggal_pak=tanggal_pak,
                user_id=user_pak if user_pak is not None else user_id,
            )
            messages.warning(
                request,
                "Data PAK sudah ada. Tolong gunakan fungsi edit.",
            )
            return redirect("profile")
        except AngkaKredit.DoesNotExist:
            AngkaKredit.objects.create(
                nomor_pak=nomor_pak,
                tanggal_pak=tanggal_pak,
                nilai_pak=nilai_pak,
                masa_penilaian_pak=masa_penilaian_pak,
                file_pak=fs.save(f"media/angkaKredit/{new_file_name}", file_pak),
                user_id=user_pak if user_pak is not None else user_id,
            )
            messages.success(request, "Data PAK berhasil dibuat.")

        return redirect("administrasi" if user_pak is not None else "profile")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def edit_pak(request, id_pak):
    user_id = request.user.id

    if request.method == "POST" and request.FILES["file_pak"]:
        nomor_pak = request.POST["nomor_pak"]
        tanggal_pak = request.POST["tanggal_pak"]
        nilai_pak = request.POST["nilai_pak"]
        masa_penilaian_pak = request.POST["masa_penilaian_pak"]
        file_pak = request.FILES["file_pak"]

        # Menggunakan get untuk mendapatkan nilai, jika ada
        user_pak = request.POST.get("user")

        # Validasi tipe file
        if not file_pak.name.endswith(".pdf"):
            messages.error(request, "File harus berformat PDF.")
            return redirect("profile")

        # Validasi ukuran file
        max_size = 1000 * 1024  # 1000KB
        if file_pak.size > max_size:
            messages.error(request, "Ukuran file melebihi 1000KB.")
            return redirect("profile")

        # Dapatkan username berdasarkan user_pak (jika ada) atau user saat ini
        if user_pak is not None:
            # Cari objek User berdasarkan id_user
            username = User.objects.get(id=user_pak).username
        else:
            # Dapatkan username dari pengguna yang saat ini masuk
            username = request.user.username

        # Ubah nama file yang akan diunggah
        new_file_name = (
            f"{username}_pak_{nomor_pak}{file_pak.name[file_pak.name.rfind('.'):]}"
        )

        fs = FileSystemStorage()

        try:
            pak = get_object_or_404(
                AngkaKredit,
                id_pak=id_pak,
                user=user_pak if user_pak is not None else user_id,
            )
            # Hapus file dan data PAK sebelumnya
            if pak.file_pak:
                old_file_path = pak.file_pak.path
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)

            # Update data PAK
            pak.nomor_pak = nomor_pak
            pak.tanggal_pak = tanggal_pak
            pak.nilai_pak = nilai_pak
            pak.masa_penilaian_pak = masa_penilaian_pak
            pak.file_pak = fs.save(f"media/angkaKredit/{new_file_name}", file_pak)
            pak.save()
            messages.success(request, "Data PAK berhasil diubah.")
        except AngkaKredit.DoesNotExist:
            messages.warning(
                request, "Data PAK tidak ditemukan. Tolong gunakan fungsi buat."
            )

        return redirect("administrasi" if user_pak is not None else "profile")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def delete_pak(request, id_pak):
    pak = get_object_or_404(AngkaKredit, id_pak=id_pak)

    try:
        # Hapus file dan data PAK
        if pak.file_pak:
            old_file_path = pak.file_pak.path
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        pak.delete()
        messages.success(request, "Data PAK berhasil dihapus.")
    except AngkaKredit.DoesNotExist:
        messages.warning(request, "Data PAK tidak ditemukan.")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def download_file_pak(request, id_pak):
    pak = get_object_or_404(AngkaKredit, id_pak=id_pak)

    # Pastikan pengguna yang mengakses adalah pemilik data File
    if pak.user != request.user:
        messages.error(request, "Anda tidak memiliki izin untuk mengunduh File ini.")
        return redirect("profile")

    # Cek apakah file_pak ada atau tidak
    if not pak.file_pak:
        raise Http404("File SK tidak ditemukan")

    # Dapatkan path lengkap ke file PAK
    file_path = pak.file_pak.path

    # Buka file sebagai binary dan kirim sebagai HTTP response
    with open(file_path, "rb") as file:
        response = HttpResponse(
            file.read(), content_type="application/pdf"
        )  # Sesuaikan content_type sesuai tipe file

        # Mengatur header untuk attachment agar file dapat diunduh
        response["Content-Disposition"] = f"attachment; filename={pak.file_pak.name}"

        return response


##### DIKLAT #####
@login_required
def add_diklat(request):
    user_id = request.user.id

    if request.method == "POST" and request.FILES["file_sertifikat_diklat"]:
        nama_diklat = request.POST["nama_diklat"]
        tanggal_mulai_diklat = request.POST["tanggal_mulai_diklat"]
        tanggal_selesai_diklat = request.POST["tanggal_selesai_diklat"]
        nomor_sertifikat_diklat = request.POST["nomor_sertifikat_diklat"]
        tanggal_sertifikat_diklat = request.POST["tanggal_sertifikat_diklat"]
        file_sertifikat_diklat = request.FILES["file_sertifikat_diklat"]

        # Menggunakan get untuk mendapatkan nilai, jika ada
        user_diklat = request.POST.get("user")

        # Validasi tipe file
        if not file_sertifikat_diklat.name.endswith(".pdf"):
            messages.error(request, "File harus berformat PDF.")
            return redirect("profile")

        # Validasi ukuran file
        max_size = 1000 * 1024  # 1000KB
        if file_sertifikat_diklat.size > max_size:
            messages.error(request, "Ukuran file melebihi 1000KB.")
            return redirect("profile")

        # Dapatkan username berdasarkan user_diklat (jika ada) atau user saat ini
        if user_diklat is not None:
            # Cari objek User berdasarkan id_user
            username = User.objects.get(id=user_diklat).username
        else:
            # Dapatkan username dari pengguna yang saat ini masuk
            username = request.user.username

        # Ubah nama file yang akan diunggah
        new_file_name = f"{username}_diklat_{nomor_sertifikat_diklat}{file_sertifikat_diklat.name[file_sertifikat_diklat.name.rfind('.'):]}"

        fs = FileSystemStorage()

        try:
            # Cek apakah data diklat sudah ada
            Diklat.objects.get(
                nomor_sertifikat_diklat=nomor_sertifikat_diklat,
                tanggal_sertifikat_diklat=tanggal_sertifikat_diklat,
                user_id=user_diklat if user_diklat is not None else user_id,
            )
            messages.warning(
                request,
                "Data Diklat sudah ada. Tolong gunakan fungsi edit.",
            )
            return redirect("profile")
        except Diklat.DoesNotExist:
            Diklat.objects.create(
                nama_diklat=nama_diklat,
                tanggal_mulai_diklat=tanggal_mulai_diklat,
                tanggal_selesai_diklat=tanggal_selesai_diklat,
                nomor_sertifikat_diklat=nomor_sertifikat_diklat,
                tanggal_sertifikat_diklat=tanggal_sertifikat_diklat,
                file_sertifikat_diklat=fs.save(
                    f"media/diklat/{new_file_name}", file_sertifikat_diklat
                ),
                user_id=user_diklat if user_diklat is not None else user_id,
            )
            messages.success(request, "Data Diklat berhasil dibuat.")

        return redirect("administrasi" if user_diklat is not None else "profile")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def edit_diklat(request, id_diklat):
    user_id = request.user.id

    if request.method == "POST" and request.FILES["file_sertifikat_diklat"]:
        nama_diklat = request.POST["nama_diklat"]
        tanggal_mulai_diklat = request.POST["tanggal_mulai_diklat"]
        tanggal_selesai_diklat = request.POST["tanggal_selesai_diklat"]
        nomor_sertifikat_diklat = request.POST["nomor_sertifikat_diklat"]
        tanggal_sertifikat_diklat = request.POST["tanggal_sertifikat_diklat"]
        file_sertifikat_diklat = request.FILES["file_sertifikat_diklat"]

        # Menggunakan get untuk mendapatkan nilai, jika ada
        user_diklat = request.POST.get("user")

        # Validasi tipe file
        if not file_sertifikat_diklat.name.endswith(".pdf"):
            messages.error(request, "File harus berformat PDF.")
            return redirect("profile")

        # Validasi ukuran file
        max_size = 1000 * 1024  # 1000KB
        if file_sertifikat_diklat.size > max_size:
            messages.error(request, "Ukuran file melebihi 1000KB.")
            return redirect("profile")

        # Dapatkan username berdasarkan user_diklat (jika ada) atau user saat ini
        if user_diklat is not None:
            # Cari objek User berdasarkan id_user
            username = User.objects.get(id=user_diklat).username
        else:
            # Dapatkan username dari pengguna yang saat ini masuk
            username = request.user.username

        # Ubah nama file yang akan diunggah
        new_file_name = f"{username}_diklat_{nomor_sertifikat_diklat}{file_sertifikat_diklat.name[file_sertifikat_diklat.name.rfind('.'):]}"

        fs = FileSystemStorage()

        try:
            diklat = get_object_or_404(
                Diklat,
                id_diklat=id_diklat,
                user=user_diklat if user_diklat is not None else user_id,
            )
            # Hapus file dan data Diklat sebelumnya
            if diklat.file_sertifikat_diklat:
                old_file_path = diklat.file_sertifikat_diklat.path
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)

            # Update data Diklat
            diklat.nama_diklat = nama_diklat
            diklat.tanggal_mulai_diklat = tanggal_mulai_diklat
            diklat.tanggal_selesai_diklat = tanggal_selesai_diklat
            diklat.nomor_sertifikat_diklat = nomor_sertifikat_diklat
            diklat.tanggal_sertifikat_diklat = tanggal_sertifikat_diklat
            diklat.file_sertifikat_diklat = fs.save(
                f"media/diklat/{new_file_name}", file_sertifikat_diklat
            )
            diklat.save()
            messages.success(request, "Data Diklat berhasil diubah.")
        except Diklat.DoesNotExist:
            messages.warning(
                request, "Data Diklat tidak ditemukan. Tolong gunakan fungsi buat."
            )

        return redirect("administrasi" if user_diklat is not None else "profile")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def delete_diklat(request, id_diklat):
    diklat = get_object_or_404(Diklat, id_diklat=id_diklat)

    try:
        # Hapus file dan data Diklat
        if diklat.file_sertifikat_diklat:
            old_file_path = diklat.file_sertifikat_diklat.path
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        diklat.delete()
        messages.success(request, "Data Diklat berhasil dihapus.")
    except Diklat.DoesNotExist:
        messages.warning(request, "Data Diklat tidak ditemukan.")

    if request.user.is_staff or request.user.is_superuser:
        # Redirect admin to administrasi.html
        return redirect("administrasi")
    else:
        # Redirect other users to profile.html
        return redirect("profile")


@login_required
def download_file_diklat(request, id_diklat):
    diklat = get_object_or_404(Diklat, id_diklat=id_diklat)

    # Pastikan pengguna yang mengakses adalah pemilik data File
    if diklat.user != request.user:
        messages.error(request, "Anda tidak memiliki izin untuk mengunduh File ini.")
        return redirect("profile")

    # Cek apakah file_sertifikat_diklat ada atau tidak
    if not diklat.file_sertifikat_diklat:
        raise Http404("File Sertifikat tidak ditemukan")

    # Dapatkan path lengkap ke file Diklat
    file_path = diklat.file_sertifikat_diklat.path

    # Buka file sebagai binary dan kirim sebagai HTTP response
    with open(file_path, "rb") as file:
        response = HttpResponse(
            file.read(), content_type="application/pdf"
        )  # Sesuaikan content_type sesuai tipe file

        # Mengatur header untuk attachment agar file dapat diunduh
        response[
            "Content-Disposition"
        ] = f"attachment; filename={diklat.file_sertifikat_diklat.name}"

        return response


##### USER #####
@login_required
@user_passes_test(lambda user: user.is_superuser or user.is_staff)
def add_user(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        email = request.POST["email"]

        if not re.match("^[a-zA-Z0-9]+$", username):
            messages.error(
                request, "Username hanya boleh terdiri dari huruf dan angka."
            )
            return redirect("administrasi")

        try:
            User.objects.get(
                username=username,
                password=password,
                email=email,
            )
            messages.warning(
                request,
                "Data User sudah ada. Tolong gunakan fungsi edit.",
            )
            return redirect("administrasi")
        except User.DoesNotExist:
            # Buat user baru
            User.objects.create_user(username=username, password=password, email=email)
            messages.success(request, "Data User berhasil dibuat.")
            return redirect("administrasi")

    return render(request, "pegawai/administrasi.html")


@login_required
@user_passes_test(lambda user: user.is_superuser or user.is_staff)
def edit_user(request, user_id):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]
        email = request.POST["email"]

        try:
            user = get_object_or_404(User, id=user_id)

            # Update data user
            user.username = username
            user.set_password(
                password
            )  # Gunakan set_password untuk mengatur password yang benar
            user.email = email
            user.save()
            messages.success(request, "Data User berhasil diubah.")
        except User.DoesNotExist:
            messages.warning(
                request, "Data User tidak ditemukan. Tolong gunakan fungsi buat."
            )

        return redirect("administrasi")

    return render(request, "pegawai/administrasi.html")


@login_required
@user_passes_test(lambda user: user.is_superuser or user.is_staff)
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)

    try:
        user.delete()
        messages.success(request, "Data User berhasil dihapus.")
    except User.DoesNotExist:
        messages.warning(request, "Data User tidak ditemukan.")

    return redirect("administrasi")


@login_required
@user_passes_test(lambda user: user.is_superuser or user.is_staff)
def export_to_excel(request):
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = "attachment; filename=exported_data.xlsx"

    workbook = openpyxl.Workbook()

    # Ekspor Data Pegawai
    pegawai_worksheet = workbook.active
    pegawai_worksheet.title = "Pegawai"
    pegawai_header = [
        "NIP Pegawai",
        "Nama Pegawai",
        "Tempat Lahir Pegawai",
        "Tanggal Lahir Pegawai",
        "Jenis Kelamin Pegawai",
        "Surel Pegawai",
        "Telepon Pegawai",
    ]
    pegawai_worksheet.append(pegawai_header)
    pegawai_queryset = Pegawai.objects.all().order_by("nip_pegawai")
    for obj in pegawai_queryset:
        row = [
            str(obj.nip_pegawai),
            obj.nama_pegawai,
            obj.tempat_lahir_pegawai,
            obj.tanggal_lahir_pegawai,
            obj.jenis_kelamin_pegawai,
            obj.surel_pegawai,
            obj.telepon_pegawai,
        ]
        pegawai_worksheet.append(row)

    # Ekspor Data Jabatan
    jabatan_worksheet = workbook.create_sheet(title="Jabatan")
    jabatan_header = [
        "NIP Pegawai",
        "Nama Pegawai",
        "Nama Jabatan",
        "Nomor SK Jabatan",
        "Tanggal SK Jabatan",
        "TMT Jabatan",
    ]
    jabatan_worksheet.append(jabatan_header)
    jabatan_queryset = Jabatan.objects.all().order_by("user__pegawai__nip_pegawai")
    for obj in jabatan_queryset:
        pegawai = Pegawai.objects.get(user=obj.user)
        row = [
            str(pegawai.nip_pegawai),
            pegawai.nama_pegawai,
            obj.get_nama_jabatan_display(),
            obj.nomor_sk_jabatan,
            obj.tanggal_sk_jabatan,
            obj.tmt_jabatan,
        ]
        jabatan_worksheet.append(row)

    # Ekspor Data Pendidikan
    pendidikan_worksheet = workbook.create_sheet(title="Pendidikan")
    pendidikan_header = [
        "NIP Pegawai",
        "Nama Pegawai",
        "Tingkat Pendidikan",
        "Lembaga Pendidikan",
        "Fakultas Pendidikan",
        "Jurusan Pendidikan",
        "Gelar Depan Pendidikan",
        "Gelar Belakang Pendidikan",
        "Nomor Seri Ijazah Pendidikan",
        "Tanggal Terbit Ijazah Pendidikan",
    ]
    pendidikan_worksheet.append(pendidikan_header)
    pendidikan_queryset = Pendidikan.objects.all().order_by(
        "user__pegawai__nip_pegawai"
    )
    for obj in pendidikan_queryset:
        pegawai = Pegawai.objects.get(user=obj.user)
        row = [
            str(pegawai.nip_pegawai),
            pegawai.nama_pegawai,
            obj.tingkat_pendidikan,
            obj.lembaga_pendidikan,
            obj.fakultas_pendidikan,
            obj.jurusan_pendidikan,
            obj.gelar_depan_pendidikan,
            obj.gelar_belakang_pendidikan,
            obj.nomor_seri_ijazah_pendidikan,
            obj.tanggal_terbit_ijazah_pendidikan,
        ]
        pendidikan_worksheet.append(row)

    # Ekspor Data Pangkat
    pangkat_worksheet = workbook.create_sheet(title="Pangkat")
    pangkat_header = [
        "NIP Pegawai",
        "Nama Pegawai",
        "Nama Pangkat",
        "Nomor SK Pangkat",
        "Tanggal SK Pangkat",
        "TMT Pangkat",
    ]
    pangkat_worksheet.append(pangkat_header)
    pangkat_queryset = Pangkat.objects.all().order_by("user__pegawai__nip_pegawai")
    for obj in pangkat_queryset:
        pegawai = Pegawai.objects.get(user=obj.user)
        row = [
            str(pegawai.nip_pegawai),
            pegawai.nama_pegawai,
            obj.nama_pangkat,
            obj.nomor_sk_pangkat,
            obj.tanggal_sk_pangkat,
            obj.tmt_pangkat,
        ]
        pangkat_worksheet.append(row)

    # Ekspor Data Diklat
    diklat_worksheet = workbook.create_sheet(title="Diklat")
    diklat_header = [
        "NIP Pegawai",
        "Nama Pegawai",
        "Nama Diklat",
        "Tanggal Mulai Diklat",
        "Tanggal Selesai Diklat",
        "Nomor Sertifikat Diklat",
        "Tanggal Sertifikat Diklat",
    ]
    diklat_worksheet.append(diklat_header)
    diklat_queryset = Diklat.objects.all().order_by("user__pegawai__nip_pegawai")
    for obj in diklat_queryset:
        pegawai = Pegawai.objects.get(user=obj.user)
        row = [
            str(pegawai.nip_pegawai),
            pegawai.nama_pegawai,
            obj.nama_diklat,
            obj.tanggal_mulai_diklat,
            obj.tanggal_selesai_diklat,
            obj.nomor_sertifikat_diklat,
            obj.tanggal_sertifikat_diklat,
        ]
        diklat_worksheet.append(row)

    # Ekspor Data Angka Kredit
    angka_kredit_worksheet = workbook.create_sheet(title="AngkaKredit")
    angka_kredit_header = [
        "NIP Pegawai",
        "Nama Pegawai",
        "Nomor Penilaian Angka Kredit",
        "Tanggal Penilaian Angka Kredit",
        "Nilai Penilaian Angka Kredit",
        "Masa Penilaian Angka Kredit",
    ]
    angka_kredit_worksheet.append(angka_kredit_header)
    angka_kredit_queryset = AngkaKredit.objects.all().order_by(
        "user__pegawai__nip_pegawai"
    )
    for obj in angka_kredit_queryset:
        pegawai = Pegawai.objects.get(user=obj.user)
        row = [
            str(pegawai.nip_pegawai),
            pegawai.nama_pegawai,
            obj.nomor_pak,
            obj.tanggal_pak,
            obj.nilai_pak,
            obj.masa_penilaian_pak,
        ]
        angka_kredit_worksheet.append(row)

    workbook.save(response)
    return response


@login_required
@user_passes_test(lambda user: user.is_superuser or user.is_staff)
def import_from_excel(request):
    if request.method == "POST" and request.FILES.get("file_import_data"):
        uploaded_file = request.FILES["file_import_data"]
        try:
            workbook = openpyxl.load_workbook(uploaded_file)

            # Memeriksa dan menangani lembar kerja "Users"
            if "Users" in workbook.sheetnames:
                users_sheet = workbook["Users"]
                if users_sheet.max_row > 1:
                    for row in users_sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 3:
                            username = row[0]
                            password = row[1]
                            email = row[2]
                            try:
                                user = User.objects.get(username=username)
                                # Update existing user's information
                                user.set_password(password)
                                user.email = email
                                user.save()
                            except ObjectDoesNotExist:
                                # Create a new user
                                user = User.objects.create_user(
                                    username=username, password=password, email=email
                                )
                        else:
                            messages.error(
                                request, "Data tidak lengkap pada lembar kerja Users"
                            )
                            return redirect("administrasi")
                else:
                    messages.error(request, "Lembar kerja Users kosong")
                    return redirect("administrasi")
            else:
                messages.error(request, "Lembar kerja Users tidak ditemukan")
                return redirect("administrasi")

            # Memeriksa dan menangani lembar kerja "Pegawai"
            if "Pegawai" in workbook.sheetnames:
                pegawai_sheet = workbook["Pegawai"]
                if pegawai_sheet.max_row > 1:
                    for row in pegawai_sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 8:
                            username = row[7]
                            try:
                                user = User.objects.get(username=username)
                                pegawai = Pegawai(
                                    nip_pegawai=row[0],
                                    nama_pegawai=row[1],
                                    tempat_lahir_pegawai=row[2],
                                    tanggal_lahir_pegawai=row[3],
                                    jenis_kelamin_pegawai=row[4],
                                    surel_pegawai=row[5],
                                    telepon_pegawai=row[6],
                                    user=user,
                                )
                                pegawai.save()
                            except User.DoesNotExist:
                                # Handle the case when the user does not exist
                                messages.error(
                                    request,
                                    "Pengguna dengan username '{username}' tidak ditemukan",
                                )
                                return redirect("administrasi")
                            except Exception as e:
                                messages.error(
                                    request,
                                    "Terjadi kesalahan saat mengimpor Pegawai: "
                                    + {str(e)},
                                )
                                return redirect("administrasi")
                        else:
                            messages.error(
                                request, "Data tidak lengkap pada lembar kerja Pegawai"
                            )
                            return redirect("administrasi")
                else:
                    messages.error(request, "Lembar kerja Pegawai kosong")
                    return redirect("administrasi")
            else:
                messages.error(request, "Lembar kerja Pegawai tidak ditemukan")
                return redirect("administrasi")

            # Memeriksa dan menangani lembar kerja "Jabatan"
            if "Jabatan" in workbook.sheetnames:
                jabatan_sheet = workbook["Jabatan"]
                if jabatan_sheet.max_row > 1:
                    for row in jabatan_sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 5:
                            username = row[4]
                            try:
                                user = User.objects.get(username=username)
                                jabatan = Jabatan(
                                    nama_jabatan=row[0],
                                    nomor_sk_jabatan=row[1],
                                    tanggal_sk_jabatan=row[2],
                                    tmt_jabatan=row[3],
                                    user=user,
                                )
                                jabatan.save()
                            except User.DoesNotExist:
                                # Handle the case when the user does not exist
                                messages.error(
                                    request,
                                    "Pengguna dengan username '{username}' tidak ditemukan",
                                )
                                return redirect("administrasi")
                            except Exception as e:
                                messages.error(
                                    request,
                                    "Terjadi kesalahan saat mengimpor Jabatan: "
                                    + {str(e)},
                                )
                                return redirect("administrasi")
                        else:
                            messages.error(
                                request, "Data tidak lengkap pada lembar kerja Jabatan"
                            )
                            return redirect("administrasi")
                else:
                    messages.error(request, "Lembar kerja Jabatan kosong")
                    return redirect("administrasi")
            else:
                messages.error(request, "Lembar kerja Jabatan tidak ditemukan")
                return redirect("administrasi")

            # Memeriksa dan menangani lembar kerja "Pendidikan"
            if "Pendidikan" in workbook.sheetnames:
                pendidikan_sheet = workbook["Pendidikan"]
                if pendidikan_sheet.max_row > 1:
                    for row in pendidikan_sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 9:
                            username = row[8]
                            try:
                                user = User.objects.get(username=username)
                                pendidikan = Pendidikan(
                                    tingkat_pendidikan=row[0],
                                    lembaga_pendidikan=row[1],
                                    fakultas_pendidikan=row[2],
                                    jurusan_pendidikan=row[3],
                                    gelar_depan_pendidikan=row[4],
                                    gelar_belakang_pendidikan=row[5],
                                    nomor_seri_ijazah_pendidikan=row[6],
                                    tanggal_terbit_ijazah_pendidikan=row[7],
                                    user=user,
                                )
                                pendidikan.save()
                            except User.DoesNotExist:
                                # Handle the case when the user does not exist
                                messages.error(
                                    request,
                                    "Pengguna dengan username '{username}' tidak ditemukan",
                                )
                                return redirect("administrasi")
                            except Exception as e:
                                messages.error(
                                    request,
                                    "Terjadi kesalahan saat mengimpor Pendidikan: "
                                    + {str(e)},
                                )
                                return redirect("administrasi")
                        else:
                            messages.error(
                                request,
                                "Data tidak lengkap pada lembar kerja Pendidikan",
                            )
                            return redirect("administrasi")
                else:
                    messages.error(request, "Lembar kerja Pendidikan kosong")
                    return redirect("administrasi")
            else:
                messages.error(request, "Lembar kerja Pendidikan tidak ditemukan")
                return redirect("administrasi")

            # Memeriksa dan menangani lembar kerja "Pangkat"
            if "Pangkat" in workbook.sheetnames:
                pangkat_sheet = workbook["Pangkat"]
                if pangkat_sheet.max_row > 1:
                    for row in pangkat_sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 5:
                            username = row[4]
                            try:
                                user = User.objects.get(username=username)
                                pangkat = Pangkat(
                                    nama_pangkat=row[0],
                                    nomor_sk_pangkat=row[1],
                                    tanggal_sk_pangkat=row[2],
                                    tmt_pangkat=row[3],
                                    user=user,
                                )
                                pangkat.save()
                            except User.DoesNotExist:
                                # Handle the case when the user does not exist
                                messages.error(
                                    request,
                                    "Pengguna dengan username '{username}' tidak ditemukan",
                                )
                                return redirect("administrasi")
                            except Exception as e:
                                messages.error(
                                    request,
                                    "Terjadi kesalahan saat mengimpor Pangkat: "
                                    + {str(e)},
                                )
                                return redirect("administrasi")
                        else:
                            messages.error(
                                request,
                                "Data tidak lengkap pada lembar kerja Pangkat",
                            )
                            return redirect("administrasi")
                else:
                    messages.error(request, "Lembar kerja Pangkat kosong")
                    return redirect("administrasi")
            else:
                messages.error(request, "Lembar kerja Pangkat tidak ditemukan")
                return redirect("administrasi")

            # Memeriksa dan menangani lembar kerja "Diklat"
            if "Diklat" in workbook.sheetnames:
                diklat_sheet = workbook["Diklat"]
                if diklat_sheet.max_row > 1:
                    for row in diklat_sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 6:
                            username = row[5]
                            try:
                                user = User.objects.get(username=username)
                                diklat = Diklat(
                                    nama_diklat=row[0],
                                    tanggal_mulai_diklat=row[1],
                                    tanggal_selesai_diklat=row[2],
                                    nomor_sertifikat_diklat=row[3],
                                    tanggal_sertifikat_diklat=row[4],
                                    user=user,
                                )
                                diklat.save()
                            except User.DoesNotExist:
                                # Handle the case when the user does not exist
                                messages.error(
                                    request,
                                    "Pengguna dengan username '{username}' tidak ditemukan",
                                )
                                return redirect("administrasi")
                            except Exception as e:
                                messages.error(
                                    request,
                                    "Terjadi kesalahan saat mengimpor Diklat: "
                                    + {str(e)},
                                )
                                return redirect("administrasi")
                        else:
                            messages.error(
                                request, "Data tidak lengkap pada lembar kerja Diklat"
                            )
                            return redirect("administrasi")
                else:
                    messages.error(request, "Lembar kerja Diklat kosong")
                    return redirect("administrasi")
            else:
                messages.error(request, "Lembar kerja Diklat tidak ditemukan")
                return redirect("administrasi")

            # Memeriksa dan menangani lembar kerja "AngkaKredit"
            if "AngkaKredit" in workbook.sheetnames:
                angka_kredit_sheet = workbook["AngkaKredit"]
                if angka_kredit_sheet.max_row > 1:
                    for row in angka_kredit_sheet.iter_rows(
                        min_row=2, values_only=True
                    ):
                        if len(row) >= 5:
                            username = row[4]
                            try:
                                user = User.objects.get(username=username)
                                angka_kredit = AngkaKredit(
                                    nomor_pak=row[0],
                                    tanggal_pak=row[1],
                                    nilai_pak=row[2],
                                    masa_penilaian_pak=row[3],
                                    user=user,
                                )
                                angka_kredit.save()
                            except User.DoesNotExist:
                                # Handle the case when the user does not exist
                                messages.error(
                                    request,
                                    "Pengguna dengan username '{username}' tidak ditemukan",
                                )
                                return redirect("administrasi")
                            except Exception as e:
                                messages.error(
                                    request,
                                    "Terjadi kesalahan saat mengimpor Penilaian Angka Kredit: "
                                    + {str(e)},
                                )
                                return redirect("administrasi")
                        else:
                            messages.error(
                                request,
                                "Data tidak lengkap pada lembar kerja Penilaian Angka Kredit",
                            )
                            return redirect("administrasi")
                else:
                    messages.error(
                        request, "Lembar kerja Penilaian Angka Kredit kosong"
                    )
                    return redirect("administrasi")
            else:
                messages.error(
                    request, "Lembar kerja Penilaian Angka Kredit tidak ditemukan"
                )
                return redirect("administrasi")

            # Ganti dengan halaman yang sesuai setelah mengimpor
            messages.success(request, "Anda berhasil Import data dari File Excel")
            return redirect("administrasi")
        except Exception as e:
            messages.error(request, "Terjadi kesalahan: " + {str(e)})
            return redirect("administrasi")

    # Ganti dengan halaman yang sesuai untuk mengimpor
    return render(request, "import_page.html")


@login_required
@user_passes_test(lambda user: user.is_superuser or user.is_staff)
def administrasi(request):
    # Mendapatkan username user yang sedang login
    username = request.user.username

    ##### USER #####
    try:
        # Mengambil data tabel pegawai
        dataUser = User.objects.all()
    except User.DoesNotExist:
        # Jika data tabel pegawai tidak ada maka return Null/None
        dataUser = None

    ##### PEGAWAI #####
    try:
        # Mengambil data tabel pegawai
        pegawai = Pegawai.objects.all()
    except Pegawai.DoesNotExist:
        # Jika data tabel pegawai tidak ada maka return Null/None
        pegawai = None

    ##### PENDIDIKAN #####
    try:
        # Mengambil data tabel pendidikan
        pendidikan = Pendidikan.objects.all().order_by(
            "-tanggal_terbit_ijazah_pendidikan"
        )
        if pendidikan.exists():
            pendidikan_terakhir = pendidikan[0]
        else:
            pendidikan_terakhir = None
    except Pendidikan.DoesNotExist:
        # Jika data tabel pendidikan tidak ada maka return Null/None
        pendidikan = None
        pendidikan_terakhir = None

    # Inisialisasi file_exists_pendidikan sebagai False
    file_exists_pendidikan = False

    if pendidikan.exists():
        # Ubah format tanggal terbit ijazah pendidikan menjadi "YYYY/MM/DD"
        for pend in pendidikan:
            if pend and pend.file_ijazah_pendidikan:
                file_exists_pendidikan = default_storage.exists(
                    pend.file_ijazah_pendidikan.name
                )
            else:
                # Keluar dari loop jika file_exists_pendidikan telah ditemukan
                break

    # Mengambil data pilihan tingkat pendidikan
    pilihan_tingkat_pendidikan = Pendidikan.TingkatPendidikan.choices

    ##### JABATAN #####
    try:
        # Mengambil data tabel jabatan
        jabatan = Jabatan.objects.all().order_by("-tanggal_sk_jabatan")
        if jabatan.exists():
            jabatan_terakhir = jabatan[0]
        else:
            jabatan_terakhir = None
    except Jabatan.DoesNotExist:
        # Jika data tabel jabatan tidak ada maka return Null/None
        jabatan = None
        jabatan_terakhir = None

    # Inisialisasi file_exists_jabatan sebagai False
    file_exists_jabatan = False

    if jabatan.exists():
        # Ubah format tanggal sk jabatan menjadi "YYYY/MM/DD"
        for jab in jabatan:
            if jab and jab.file_sk_jabatan:
                file_exists_jabatan = default_storage.exists(jab.file_sk_jabatan.name)
            else:
                # Keluar dari loop jika file_exists_jabatan telah ditemukan
                break

    # Mengambil data pilihan jabatan
    pilihan_jabatan = Jabatan.NamaJabatan.choices

    ##### PANGKAT #####
    try:
        # Mengambil data tabel pangkat
        pangkat = Pangkat.objects.all().order_by("-tanggal_sk_pangkat")
        if pangkat.exists():
            pangkat_terakhir = pangkat[0]
        else:
            pangkat_terakhir = None
    except Pangkat.DoesNotExist:
        # Jika data tabel pangkat tidak ada maka return Null/None
        pangkat = None
        pangkat_terakhir = None

    # Inisialisasi file_exists_pangkat sebagai False
    file_exists_pangkat = False

    if pangkat.exists():
        # Ubah format tanggal sk pangkat menjadi "YYYY/MM/DD"
        for pang in pangkat:
            if pang and pang.tanggal_sk_pangkat:
                file_exists_pangkat = default_storage.exists(pang.file_sk_pangkat.name)
            else:
                # Keluar dari loop jika file_exists_pangkat telah ditemukan
                break

    # Mengambil data pilihan pangkat
    pilihan_pangkat = Pangkat.KodePangkat.choices

    ##### PENILAIAN ANGKA KREDIT (PAK) #####
    try:
        # Mengambil data tabel PAK
        pak = AngkaKredit.objects.all().order_by("-tanggal_pak")
        if pak.exists():
            pak_terakhir = pak[0]
        else:
            pak_terakhir = None
    except AngkaKredit.DoesNotExist:
        # Jika data tabel PAK tidak ada maka return Null/None
        pak = None
        pak_terakhir = None

    # Inisialisasi file_exists_pak sebagai False
    file_exists_pak = False

    # Mengambil data PAK untuk dimasukkan ke dalam chart
    pak_data = []

    if pak.exists():
        for ak in pak:
            pak_data.append(
                {
                    "nilai_pak": ak.nilai_pak,
                }
            )
            if ak:
                file_exists_pak = default_storage.exists(ak.file_pak.name)
            else:
                # Keluar dari loop jika file_exists_pak telah ditemukan
                break

    # Mengubah data menjadi JSON
    pak_data_json = json.dumps(pak_data)

    ##### DIKLAT #####
    try:
        # Mengambil data tabel DIKLAT
        diklat = Diklat.objects.all().order_by("-tanggal_sertifikat_diklat")
        if diklat.exists():
            diklat_terakhir = diklat[0]
        else:
            diklat_terakhir = None
    except Diklat.DoesNotExist:
        # Jika data tabel DIKLAT tidak ada maka return Null/None
        diklat = None
        diklat_terakhir = None

    # Inisialisasi file_exists_dik sebagai False
    file_exists_diklat = False

    if diklat.exists():
        # Ubah format tanggal sertifikat DIKLAT, tanggal mulai dan selesai DIKLAT menjadi "YYYY/MM/DD"
        for dik in diklat:
            if dik and dik.tanggal_sertifikat_diklat:
                file_exists_diklat = default_storage.exists(
                    dik.file_sertifikat_diklat.name
                )
            else:
                # Keluar dari loop jika file_exists_diklat telah ditemukan
                break

    context = {
        "username": username,
        "dataUser": dataUser,
        "pegawai": pegawai,
        "pendidikan": pendidikan,
        "pendidikan_terakhir": pendidikan_terakhir,
        "pilihan_tingkat_pendidikan": pilihan_tingkat_pendidikan,
        "file_exists_pendidikan": file_exists_pendidikan,
        "jabatan": jabatan,
        "jabatan_terakhir": jabatan_terakhir,
        "pilihan_jabatan": pilihan_jabatan,
        "file_exists_jabatan": file_exists_jabatan,
        "pangkat": pangkat,
        "pangkat_terakhir": pangkat_terakhir,
        "pilihan_pangkat": pilihan_pangkat,
        "file_exists_pangkat": file_exists_pangkat,
        "pak": pak,
        "pak_terakhir": pak_terakhir,
        "file_exists_pak": file_exists_pak,
        "pak_data_json": pak_data_json,  # Menambahkan data JSON ke konteks
        "diklat": diklat,
        "diklat_terakhir": diklat_terakhir,
        "file_exists_diklat": file_exists_diklat,
    }
    return render(request, "pegawai/administrasi.html", context)
